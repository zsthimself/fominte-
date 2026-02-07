import openpyxl
import csv

# 读取 Excel
wb = openpyxl.load_workbook(r'e:\Bottle1\产品上新\8件摇粒绒.xlsx')
ws = wb.active

# 写入 CSV
with open(r'e:\Bottle1\产品上新\产品数据.csv', 'w', encoding='utf-8-sig', newline='') as f:
    writer = csv.writer(f)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)

print(f" 成功导出 {ws.max_row} 行 x {ws.max_column} 列数据到 CSV")
